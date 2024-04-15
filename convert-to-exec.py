# after running pyinstaller on this file to create executable, move the exe file from dist folder to folder where pivot_table is there, then run it.


# exe will access the pivot_table.xlsx file only if it is in the same folder



from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

##same as pivot-to-report.py but import os, sys, change how input of workbook is done and how file is saved
#done to help create executable
#COMPLETE AUTOMATION, JUST ENTER MONTH AND FILE NAME OF THE PIVOT TABLE

import os
import sys

# Preparing script before we convert it to executable
application_path = os.path.dirname(sys.executable)


month = input('Input month: ')
pivot_table = input('File Name of the pivot table (include .xlsx also): ')

# Read workbook and select sheet
input_path = os.path.join(application_path, pivot_table)
wb = load_workbook(input_path)
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)  
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1)
    letter = get_column_letter(i)
    sheet.column_dimensions[letter].width = 30   #increasing column size as the numbers are long, if the column width is not enoug, then it will show as ####

    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

# Add format
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

#saving the file to an exe
output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)
