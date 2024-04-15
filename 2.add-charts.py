from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Read workbook and select sheet
wb = load_workbook('pivot_table.xlsx')          ##charts are made on pivot_table, so we load it as the wb(workbook)
sheet = wb['Report']                            ##report sheet of wb is to be used


min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column)   #1st column has min values -- it is the col with names of Male, Female 
print(max_column)   #7th column, last row has max values
print(min_row)      #1st row is filled with categories, this is the min

barchart = BarChart()       # Instantiate a barchart

# Locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,   #min col is being given as the labels male and female, so we move it to the next column
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)

#reference creates a reference from min_col to max_col x min_row to max_row, all these starting & ending columns/rows are to be specified

#basically create references to the data, removing category names here by shifting min_column
#data contains the categories also

categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,      #max_col is taken as min_column so that Male Female only come, no other column comes, starting col is ending col 
                       min_row=min_row+1,       
                       max_row=max_row)  

#created reference to only the categories of the table


# Adding data and categories
barchart.add_data(data, titles_from_data=True)   #data for barchart is taken from data reference, title taken form data only
barchart.set_categories(categories)     #categories of barchart

# Make chart
sheet.add_chart(barchart, "B12")                #we make the barchart at B12 cell
barchart.title = 'Sales by Product line'
barchart.style = 6  # chose a style for the barchart

# Save workbook
wb.save('barchart.xlsx')     #save barchart ot excel
