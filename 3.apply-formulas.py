from openpyxl import load_workbook
from openpyxl.utils import get_column_letter   #columns are stored as numbers when openpyxl used, but using get_column_letter gives column letter like excel

wb = load_workbook('barchart.xlsx')     ##access the workbook
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# format of any formula
# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'


#applying that formula in a loop for all columns
# We go to each column one by one
for i in range(min_column+1, max_column+1):  ##min_col+1 because min_col is Male, Female   #max_col+1 because range goes thru last but one values only. so max_col+1 will ensure it goes thru max_col
    letter = get_column_letter(i)       #get current column ka letter
    sheet.column_dimensions[letter].width = 30   #increasing column size as the numbers are long, if the column width is not enoug, then it will show as ####

    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'      #we sum from minrow+1 to max_row (ie) we sum from the first value in column till the last value in column
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'
            
# address of cell = {column_letter}{row_number}   

# we choose row as max_row+1 as we want to put the value right below the last value in the same column. (last row+1)

wb.save('report.xlsx')
