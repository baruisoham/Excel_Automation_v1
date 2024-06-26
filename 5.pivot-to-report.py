#integrating 2,3,4 all together so each month can be done from right here as long as pivot_table is already produced.
#inputs are month and pivot table generated for each month, rest all it will produce


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



month = 'February'

# Read workbook and select sheet
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Instantiate a barchart
barchart = BarChart()

# Locate data and categories
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)  # including headers
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)  # not including headers

# Adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Make chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1): 
    letter = get_column_letter(i)
    sheet.column_dimensions[letter].width = 30   #increasing column size as the numbers are long, if the column width is not enoug, then it will show as ####

    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

# Add format
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, italic=True , size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save(f'report_{month}.xlsx')
